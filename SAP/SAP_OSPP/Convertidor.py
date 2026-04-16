#!/usr/bin/env python3
"""
=============================================================================
UPSERT de registros OSPP hacia [QUIRAMA].[dbo].[IntegracionSAP_OSPP]
=============================================================================
Lee un archivo Excel (.xlsx) con las columnas:
    ItemCode, CardCode, Price, Currency, DiscountPercent,
    PriceListNum, AutoUpdate, SourcePrice, ValidFrom, ToFrom, Valid

Y ejecuta un UPSERT (MERGE) contra la tabla IntegracionSAP_OSPP en SQL Server.
Llave de coincidencia: ItemCode + CardCode + PriceListNum

Uso:
    python upsert_ospp.py datos.xlsx
    python upsert_ospp.py datos.xlsx --sheet "Hoja1"
    python upsert_ospp.py datos.xlsx --dry-run
=============================================================================
Dependencias:
    pip install pyodbc openpyxl
=============================================================================
"""

import pyodbc
import sys
import os
import argparse
import logging
from datetime import datetime, date
from openpyxl import load_workbook

# =============================================================================
# CONFIGURACION DE CONEXION A SQL SERVER
# =============================================================================
DB_CONFIG = {
    "server":   "20.110.82.192\SQLEXPRESS",         # Cambiar al servidor real
    "database": "QUIRAMA",
    "username": "ConexionDiscordoba",                # Cambiar al usuario real
    "password": "Srv1Discordoba",     # Cambiar a la contraseña real
    "driver":   "{ODBC Driver 17 for SQL Server}",  # Ajustar si es necesario
}

# Alternativa: usar conexion con Windows Authentication (descomentar):
# DB_CONFIG = {
#     "server":   "localhost",
#     "database": "QUIRAMA",
#     "driver":   "{ODBC Driver 17 for SQL Server}",
#     "trusted":  True,
# }

# =============================================================================
# MAPEO DE COLUMNAS DEL EXCEL
# =============================================================================
# IMPORTANTE: El orden de busqueda importa. ValidFrom/ValidTo se buscan ANTES
# que Valid para evitar que "válido desde" matchee con "válido".
# Se usa coincidencia parcial (case-insensitive).
COLUMN_MAP_ORDERED = [
    ("ItemCode",        ["itemcode", "número de artículo", "numero de articulo", "articulo"]),
    ("CardCode",        ["cardcode", "código sn", "codigo sn", "socio de negocio"]),
    ("Price",           ["price", "precio especial", "precio"]),
    ("Currency",        ["currency", "moneda del precio", "moneda"]),
    ("DiscountPercent", ["discountpercent", "discount", "descuento"]),
    ("PriceListNum",    ["pricelistnum", "listnum", "lista de precios"]),
    ("AutoUpdate",      ["autoupdate", "autoupdt", "autoupd"]),
    ("SourcePrice",     ["sourceprice", "precio fuente"]),
    ("ValidFrom",       ["validfrom", "createdate", "válido desde", "valido desde"]),
    ("ValidTo",         ["validto", "tofrom", "válido hasta", "valido hasta"]),
    ("Valid",           ["válido", "valido", "valid", "activo"]),
]

# =============================================================================
# CONFIGURACION DE LOGGING
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# =============================================================================
# TAMANO DE LOTE PARA INSERTS EN STAGING
# =============================================================================
BATCH_SIZE = 500


def get_connection():
    """Crea y retorna una conexion a SQL Server."""
    cfg = DB_CONFIG
    if cfg.get("trusted"):
        conn_str = (
            f"DRIVER={cfg['driver']};"
            f"SERVER={cfg['server']};"
            f"DATABASE={cfg['database']};"
            f"Trusted_Connection=yes;"
        )
    else:
        conn_str = (
            f"DRIVER={cfg['driver']};"
            f"SERVER={cfg['server']};"
            f"DATABASE={cfg['database']};"
            f"UID={cfg['username']};"
            f"PWD={cfg['password']};"
        )
    logger.info(f"Conectando a {cfg['server']}/{cfg['database']}...")
    conn = pyodbc.connect(conn_str)
    logger.info("Conexion establecida.")
    return conn


def safe_str(value):
    """Convierte un valor a string limpio, retorna '' si es None."""
    if value is None:
        return ""
    return str(value).strip()


def parse_date(value):
    """
    Convierte un valor de celda Excel a objeto date.
    Maneja: datetime, date, string con formatos dd/mm/yyyy, yyyy-mm-dd, etc.
    """
    if value is None:
        return None

    # openpyxl ya entrega datetime/date para celdas con formato fecha
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    # Si es string, intentar parsear
    date_str = str(value).strip()
    if not date_str:
        return None

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue

    logger.warning(f"Formato de fecha no reconocido: '{date_str}'")
    return None


def parse_price(value):
    """
    Convierte valor de celda Excel a float.
    Maneja: numeros directos, strings como '44,330.000'.
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    price_str = str(value).strip().replace(",", "")
    if not price_str:
        return 0.0
    try:
        return float(price_str)
    except ValueError:
        logger.warning(f"Precio no valido: '{value}', usando 0.0")
        return 0.0


def parse_discount(value):
    """Convierte porcentaje de descuento a float."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return 0.0


def parse_pricelist(value):
    """Convierte PriceListNum a int."""
    if value is None:
        return 1
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(str(value).strip())
    except ValueError:
        return 1


def detect_columns(header_row):
    """
    Detecta el indice de cada columna en el encabezado del Excel.
    Usa busqueda ordenada: una columna ya asignada NO se reutiliza.
    Retorna un dict {campo: indice_columna} o None si no encuentra encabezado.
    """
    if not header_row:
        return None

    headers = [safe_str(cell).lower() for cell in header_row]
    col_indices = {}
    used_indices = set()  # Columnas ya asignadas, no reutilizar

    for field, keywords in COLUMN_MAP_ORDERED:
        found = False
        for idx, header in enumerate(headers):
            if idx in used_indices:
                continue  # Ya fue asignada a otro campo
            if any(kw in header for kw in keywords):
                col_indices[field] = idx
                used_indices.add(idx)
                found = True
                break
        if not found:
            col_indices[field] = None

    # Verificar columnas clave
    required = ["ItemCode", "CardCode", "Price"]
    missing = [f for f in required if col_indices.get(f) is None]
    if missing:
        logger.warning(f"Columnas requeridas no encontradas: {missing}")
        logger.info(f"Encabezados detectados: {headers}")
        return None

    found_cols = {k: v for k, v in col_indices.items() if v is not None}
    logger.info(f"Columnas mapeadas: {found_cols}")

    # Diagnostico: mostrar que encabezado mapeo a cada campo
    for field, idx in col_indices.items():
        if idx is not None and idx < len(headers):
            logger.info(f"  {field:20s} -> columna {idx} ('{headers[idx]}')")
        else:
            logger.info(f"  {field:20s} -> NO ENCONTRADA")

    return col_indices


def get_cell(row, index, default=None):
    """Obtiene el valor de una celda de forma segura."""
    if index is None or index >= len(row):
        return default
    return row[index]


def parse_xlsx_row(row, col_map, row_num=None):
    """Parsea una fila del Excel a un diccionario de registro."""
    item_code = safe_str(get_cell(row, col_map.get("ItemCode")))
    card_code = safe_str(get_cell(row, col_map.get("CardCode")))

    if not item_code or not card_code:
        return None

    currency_val = safe_str(get_cell(row, col_map.get("Currency")))
    auto_update_val = safe_str(get_cell(row, col_map.get("AutoUpdate")))
    source_price_val = safe_str(get_cell(row, col_map.get("SourcePrice")))
    valid_val = safe_str(get_cell(row, col_map.get("Valid")))

    # --- Validar y truncar campos de 1 caracter ---
    # AutoUpdate debe ser 'Y' o 'N' (1 caracter)
    if len(auto_update_val) > 1:
        logger.warning(
            f"  Fila {row_num}: AutoUpdate='{auto_update_val}' tiene {len(auto_update_val)} chars, "
            f"truncando a '{auto_update_val[0]}'"
        )
        auto_update_val = auto_update_val[0]

    # Valid debe ser 'Y' o 'N' (1 caracter)
    if len(valid_val) > 1:
        logger.warning(
            f"  Fila {row_num}: Valid='{valid_val}' tiene {len(valid_val)} chars, "
            f"truncando a '{valid_val[0]}'"
        )
        valid_val = valid_val[0]

    # Currency max 10 chars
    if len(currency_val) > 10:
        currency_val = currency_val[:10]

    return {
        "ItemCode":        item_code[:100],
        "CardCode":        card_code[:100],
        "Price":           parse_price(get_cell(row, col_map.get("Price"))),
        "Currency":        currency_val if currency_val else "$",
        "DiscountPercent": parse_discount(get_cell(row, col_map.get("DiscountPercent"))),
        "PriceListNum":    parse_pricelist(get_cell(row, col_map.get("PriceListNum"))),
        "AutoUpdate":      auto_update_val if auto_update_val else "Y",
        "SourcePrice":     source_price_val[:100] if source_price_val else "",
        "ValidFrom":       parse_date(get_cell(row, col_map.get("ValidFrom"))),
        "ValidTo":         parse_date(get_cell(row, col_map.get("ValidTo"))),
        "Valid":           valid_val if valid_val else "Y",
    }


def read_xlsx(filepath, sheet_name=None):
    """
    Lee un archivo .xlsx y retorna una lista de diccionarios con los registros.
    Detecta columnas por encabezado; si no lo encuentra, usa posicion.
    """
    if not os.path.exists(filepath):
        logger.error(f"Archivo no encontrado: {filepath}")
        sys.exit(1)

    logger.info(f"Leyendo archivo Excel: {filepath}")
    wb = load_workbook(filepath, read_only=True, data_only=True)

    # Seleccionar hoja
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            logger.error(f"Hoja '{sheet_name}' no encontrada. Disponibles: {wb.sheetnames}")
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb.active
        logger.info(f"Usando hoja activa: '{ws.title}'")

    logger.info(f"Hojas disponibles: {wb.sheetnames}")

    records = []
    rows_iter = ws.iter_rows(values_only=True)

    # Leer primera fila para detectar encabezado
    first_row = next(rows_iter, None)
    if first_row is None:
        logger.error("Archivo Excel vacio.")
        wb.close()
        sys.exit(1)

    # Intentar detectar columnas por encabezado
    col_map = detect_columns(first_row)

    if col_map is None:
        # Sin encabezado, usar orden por posicion
        logger.info("No se detecto encabezado. Usando orden por posicion (0-10).")
        col_map = {
            "ItemCode": 0, "CardCode": 1, "Price": 2, "Currency": 3,
            "DiscountPercent": 4, "PriceListNum": 5, "AutoUpdate": 6,
            "SourcePrice": 7, "ValidFrom": 8, "ValidTo": 9, "Valid": 10,
        }
        # Primera fila es dato
        record = parse_xlsx_row(first_row, col_map, row_num=1)
        if record:
            records.append(record)

    # Leer resto de filas
    row_num = 2
    for row in rows_iter:
        if not row or all(cell is None for cell in row):
            row_num += 1
            continue

        item_val = get_cell(row, col_map.get("ItemCode"))
        if item_val is None or safe_str(item_val) == "":
            row_num += 1
            continue

        try:
            record = parse_xlsx_row(row, col_map, row_num=row_num)
            if record:
                records.append(record)
        except Exception as e:
            logger.warning(f"Error en fila {row_num}: {e}")

        row_num += 1

    wb.close()
    logger.info(f"Registros leidos: {len(records)}")

    # Diagnostico: mostrar primeros 3 registros parseados
    if records:
        logger.info("--- Diagnostico: primeros registros parseados ---")
        for i, r in enumerate(records[:3]):
            logger.info(
                f"  [{i+1}] ItemCode='{r['ItemCode']}' | CardCode='{r['CardCode']}' | "
                f"Price={r['Price']} | Currency='{r['Currency']}' | "
                f"Disc={r['DiscountPercent']} | PL={r['PriceListNum']} | "
                f"AutoUpd='{r['AutoUpdate']}' | Valid='{r['Valid']}' | "
                f"ValidFrom={r['ValidFrom']} | ValidTo={r['ValidTo']}"
            )
        logger.info("--- Fin diagnostico ---")

    return records


def upsert_records(conn, records):
    """
    Ejecuta el UPSERT (MERGE) usando tabla temporal como staging.
    """
    cursor = conn.cursor()

    try:
        # --- Paso 1: Crear tabla temporal ---
        logger.info("Creando tabla temporal #OSPP_Staging...")
        cursor.execute("""
            IF OBJECT_ID('tempdb..#OSPP_Staging') IS NOT NULL 
                DROP TABLE #OSPP_Staging;

            CREATE TABLE #OSPP_Staging (
                ItemCode        NVARCHAR(100),
                CardCode        NVARCHAR(100),
                Price           DECIMAL(18,3),
                Currency        NVARCHAR(10),
                DiscountPercent DECIMAL(18,4),
                PriceListNum    INT,
                AutoUpdate      NVARCHAR(1),
                SourcePrice     NVARCHAR(100),
                Valid           NVARCHAR(1),
                ValidFrom       DATE,
                ValidTo         DATE
            );
        """)

        # --- Paso 2: Insertar en staging por lotes ---
        insert_sql = """
            INSERT INTO #OSPP_Staging 
                (ItemCode, CardCode, Price, Currency, DiscountPercent, 
                 PriceListNum, AutoUpdate, SourcePrice, Valid, ValidFrom, ValidTo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """

        total = len(records)
        for i in range(0, total, BATCH_SIZE):
            batch = records[i : i + BATCH_SIZE]
            params = [
                (
                    r["ItemCode"],
                    r["CardCode"],
                    r["Price"],
                    r["Currency"],
                    r["DiscountPercent"],
                    r["PriceListNum"],
                    r["AutoUpdate"],
                    r["SourcePrice"],
                    r["Valid"],
                    r["ValidFrom"],
                    r["ValidTo"],
                )
                for r in batch
            ]
            cursor.executemany(insert_sql, params)
            logger.info(f"  Staging: {min(i + BATCH_SIZE, total)}/{total} registros")

        # --- Paso 3: Ejecutar MERGE ---
        logger.info("Ejecutando MERGE (UPSERT)...")
        merge_sql = """
            MERGE [dbo].[IntegracionSAP_OSPP] AS T
            USING #OSPP_Staging AS S
            ON  T.ItemCode     = S.ItemCode 
            AND T.CardCode     = S.CardCode 
            AND T.PriceListNum = S.PriceListNum

            WHEN MATCHED THEN
                UPDATE SET
                    T.Price            = S.Price,
                    T.Currency         = S.Currency,
                    T.DiscountPercent  = S.DiscountPercent,
                    T.AutoUpdate       = S.AutoUpdate,
                    T.SourcePrice      = S.SourcePrice,
                    T.Valid            = S.Valid,
                    T.ValidFrom        = S.ValidFrom,
                    T.ValidTo          = S.ValidTo,
                    T.UpdatedAt        = GETDATE()

            WHEN NOT MATCHED BY TARGET THEN
                INSERT (ItemCode, CardCode, Price, Currency, DiscountPercent, 
                        PriceListNum, AutoUpdate, SourcePrice, Valid, 
                        ValidFrom, ValidTo, CreatedAt, UpdatedAt)
                VALUES (S.ItemCode, S.CardCode, S.Price, S.Currency, S.DiscountPercent, 
                        S.PriceListNum, S.AutoUpdate, S.SourcePrice, S.Valid, 
                        S.ValidFrom, S.ValidTo, GETDATE(), GETDATE());
        """
        cursor.execute(merge_sql)
        rows_affected = cursor.rowcount
        logger.info(f"MERGE completado. Registros afectados: {rows_affected}")

        # --- Paso 4: Limpiar ---
        cursor.execute("DROP TABLE #OSPP_Staging;")
        conn.commit()
        logger.info("Transaccion confirmada (COMMIT).")

        return rows_affected

    except Exception as e:
        conn.rollback()
        logger.error(f"Error durante UPSERT: {e}")
        raise
    finally:
        cursor.close()


def main():
    parser = argparse.ArgumentParser(
        description="UPSERT de registros OSPP (.xlsx) hacia IntegracionSAP_OSPP"
    )
    parser.add_argument(
        "archivo",
        help="Ruta al archivo Excel (.xlsx)",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Nombre de la hoja del Excel (default: hoja activa)",
    )
    parser.add_argument(
        "--server", "-s",
        default=None,
        help="Servidor SQL Server (sobreescribe config interna)",
    )
    parser.add_argument(
        "--database", "-db",
        default=None,
        help="Base de datos (sobreescribe config interna)",
    )
    parser.add_argument(
        "--username", "-u",
        default=None,
        help="Usuario SQL Server",
    )
    parser.add_argument(
        "--password", "-p",
        default=None,
        help="Contraseña SQL Server",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Solo leer y parsear el Excel, no conectar a la base de datos",
    )

    args = parser.parse_args()

    # Validar extension
    if not args.archivo.lower().endswith((".xlsx", ".xlsm")):
        logger.error("El archivo debe ser .xlsx o .xlsm")
        sys.exit(1)

    # Sobreescribir config si se pasan parametros
    if args.server:
        DB_CONFIG["server"] = args.server
    if args.database:
        DB_CONFIG["database"] = args.database
    if args.username:
        DB_CONFIG["username"] = args.username
    if args.password:
        DB_CONFIG["password"] = args.password

    logger.info("=" * 60)
    logger.info("UPSERT IntegracionSAP_OSPP desde Excel - Inicio")
    logger.info("=" * 60)

    # Leer archivo Excel
    records = read_xlsx(args.archivo, sheet_name=args.sheet)

    if not records:
        logger.warning("No se encontraron registros para procesar.")
        sys.exit(0)

    # Mostrar resumen
    logger.info(f"Total registros a procesar: {len(records)}")
    logger.info(f"Ejemplo primer registro: {records[0]}")

    if args.dry_run:
        logger.info("** DRY RUN ** - No se ejecutara contra la base de datos.")
        logger.info("Primeros 5 registros:")
        for i, r in enumerate(records[:5]):
            logger.info(
                f"  [{i+1}] {r['ItemCode']} | {r['CardCode']} | "
                f"${r['Price']:,.3f} | PL:{r['PriceListNum']} | "
                f"Disc:{r['DiscountPercent']}% | Desde:{r['ValidFrom']}"
            )
        if len(records) > 5:
            logger.info(f"  ... y {len(records) - 5} registros mas.")
        sys.exit(0)

    # Conectar y ejecutar UPSERT
    conn = None
    try:
        conn = get_connection()
        rows = upsert_records(conn, records)
        logger.info("=" * 60)
        logger.info(f"Proceso finalizado. Registros afectados: {rows}")
        logger.info("=" * 60)
    except Exception as e:
        logger.error(f"Error fatal: {e}")
        sys.exit(1)
    finally:
        if conn:
            conn.close()
            logger.info("Conexion cerrada.")


if __name__ == "__main__":
    main()
